"""Add an elapsed_minutes column to session2.xlsx (in place)."""
from openpyxl import load_workbook, Workbook
from pathlib import Path
import sys
sys.stdout.reconfigure(line_buffering=True)

src = Path(r"c:\Users\stefa\Desktop\TH-Tool\session2.xlsx")
dst = src.with_name("session2_with_elapsed.xlsx")
wb_in = load_workbook(src)
ws_in = wb_in.active

rows = list(ws_in.iter_rows(min_row=2, values_only=True))
ts0 = rows[0][0]

wb = Workbook()
ws = wb.active
ws.title = "records"
ws.append(["timestamp", "elapsed_minutes", "temperature_c", "humidity_rh"])
for ts, t, rh in rows:
    elapsed = (ts - ts0).total_seconds() / 60.0
    ws.append([ts, elapsed, t, rh])

ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 14
for cell in ws["A"][1:]: cell.number_format = "yyyy-mm-dd hh:mm:ss"
for cell in ws["B"][1:]: cell.number_format = "0"           # whole minutes
for cell in ws["C"][1:] + ws["D"][1:]: cell.number_format = "0.00"

wb.save(dst)
print(f"Wrote {dst}  ({len(rows)} rows, elapsed_minutes 0 .. {(rows[-1][0]-ts0).total_seconds()/60:.0f})")
