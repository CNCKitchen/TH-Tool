"""trh.py — CLI for the LOG_TRH (CEM DT-191A / RS-191A) data logger.

Subcommands:
  info                  show device info, current settings, live reading
  fetch  PATH.csv       dump all records to CSV (one-shot per power cycle)
  set    [options]      change one or more settings
                        --led N                LED flash cycle in seconds
                        --sample N             sample rate in seconds
                        --start manual|instant start mode
                        --unit  c|f            temperature display unit
                        --name STR             logger name (max ~16 chars)
                        --time                 sync device clock to host clock
"""
import argparse, hid, time, struct, sys, datetime, csv

VID, PID = 0x2047, 0x0301
RID = 0x3F

# ---------------------------------------------------------------- transport

def open_dev():
    for d in hid.enumerate(VID, PID):
        h = hid.device()
        h.open_path(d["path"])
        return h
    sys.exit("LOG_TRH device not found (VID 0x2047 PID 0x0301).")

def make_req(payload: bytes) -> bytes:
    body = bytes([len(payload)]) + payload
    body = body + b"\x00" * (63 - len(body))
    return bytes([RID]) + body  # 64 bytes total: report-id + len + 62 data

def read_until_idle(h, idle_ms=300, max_ms=8000):
    h.set_nonblocking(True)
    out, last, t0 = [], time.time(), time.time()
    while True:
        if (time.time() - last) * 1000 >= idle_ms: break
        if (time.time() - t0)   * 1000 >= max_ms:  break
        r = h.read(64, timeout_ms=50)
        if r:
            out.append(bytes(r)); last = time.time()
    return out

def request(h, payload, idle_ms=300, max_ms=2000):
    """Send a single-byte (or multi-byte) command and return all input reports."""
    h.write(make_req(payload))
    return read_until_idle(h, idle_ms=idle_ms, max_ms=max_ms)

# ---------------------------------------------------------------- decoders

def get_meta(h):
    """op=0x04: 31-byte structured response with current settings + start time."""
    rs = request(h, bytes([0x04]))
    for r in rs:
        if r[0] == RID and r[1] == 0x1F and r[2:5] == b"\x00\x00\x04":
            app = r[2:2 + r[1]]   # 31 bytes
            return app
    raise RuntimeError(f"unexpected meta response: {[x.hex(' ') for x in rs]}")

def parse_meta(app: bytes):
    """Decode the 31-byte META structure."""
    assert len(app) == 31 and app[2] == 0x04
    return {
        "max_points":     int.from_bytes(app[3:5],  "little"),
        "record_count":   int.from_bytes(app[5:7],  "little"),
        "sample_sec":     int.from_bytes(app[7:11], "little"),  # 4-byte LE
        "led_cycle_sec":  app[11],   # NOTE: shares byte 11 with sample_sec MSB;
                                     # safe for sample_sec <= 24 days (rate fits 3B)
        "start_mode":     app[12],   # 0=Instant, 1=Manual
        "unit":           app[22],   # 0=C, 1=F  (corresponds to write[20])
        "cal_block":      app[14:22],
        "humid_high":     int.from_bytes(app[20:22], "little") / 100.0,
        "start_time":     decode_time(app[23:30]),
        "raw":            app,
    }

def decode_time(b7: bytes) -> datetime.datetime:
    year = int.from_bytes(b7[0:2], "little")
    return datetime.datetime(year, b7[2], b7[3], b7[4], b7[5], b7[6])

def encode_time(t: datetime.datetime) -> bytes:
    return t.year.to_bytes(2, "little") + bytes([t.month, t.day, t.hour, t.minute, t.second])

def parse_duration(s: str) -> int:
    """Accept '300', '5m', '1h', '4s'. Returns seconds."""
    s = s.strip().lower()
    if s.endswith("h"): return int(s[:-1]) * 3600
    if s.endswith("m"): return int(s[:-1]) * 60
    if s.endswith("s"): return int(s[:-1])
    return int(s)

def fmt_duration(sec: int) -> str:
    if sec >= 3600 and sec % 3600 == 0: return f"{sec // 3600}h"
    if sec >=   60 and sec %   60 == 0: return f"{sec // 60}m"
    return f"{sec}s"

def fmt_long_duration(sec: int) -> str:
    """Format a span like 5994000s as '69d 8h'. Drops zero leading units."""
    d, sec = divmod(sec, 86400)
    h, sec = divmod(sec, 3600)
    m, s   = divmod(sec, 60)
    parts = []
    if d: parts.append(f"{d}d")
    if h: parts.append(f"{h}h")
    if m: parts.append(f"{m}m")
    if s or not parts: parts.append(f"{s}s")
    return " ".join(parts)

def get_model(h):
    """op=0x30 response: app[0] is the opcode echo, then ASCII model."""
    rs = request(h, bytes([0x30]))
    for r in rs:
        app = r[2:2 + r[1]]
        if r[0] == RID and app and app[0] == 0x30:
            return app[1:].split(b"\x00", 1)[0].decode("ascii", "replace")
    return "?"

def get_name(h):
    """op=0x22 response: 55-byte structure with logger name as ASCII at offset 10."""
    rs = request(h, bytes([0x22]))
    for r in rs:
        if r[0] == RID and r[1] == 0x37:
            app = r[2:2 + r[1]]
            return app[10:26].split(b"\x00", 1)[0].decode("ascii", "replace")
    return "?"

def get_live(h):
    """op=0x06 response: app = [00 00 06 T_lo T_hi RH_lo RH_hi]."""
    rs = request(h, bytes([0x06]))
    for r in rs:
        app = r[2:2 + r[1]]
        if r[0] == RID and len(app) >= 7 and app[:3] == b"\x00\x00\x06":
            t  = int.from_bytes(app[3:5], "little", signed=True) / 100.0
            rh = int.from_bytes(app[5:7], "little", signed=True) / 100.0
            return t, rh
    return None, None

# ---------------------------------------------------------------- writes

def write_settings(h, meta_app: bytes, *, sample=None, led=None,
                   start_mode=None, unit=None, set_time=None):
    """Compose op=0x03 payload from current meta, applying overrides."""
    a = bytearray(meta_app)
    # The 28-byte settings payload sits at meta_app[3..30] in the read response.
    # Field positions within the WRITE payload (after opcode byte 0x03):
    payload = bytearray(29)
    payload[0]    = 0x03
    payload[1:3]  = a[3:5]    # max_points (preserve)
    payload[3:5]  = b"\x00\x00"  # record_count slot — zeroed on write
    sample_val = sample if sample is not None else int.from_bytes(a[7:11], "little")
    payload[5:9] = sample_val.to_bytes(4, "little")
    payload[9]   = led if led is not None else a[11]
    payload[10]   = start_mode if start_mode is not None else a[12]
    payload[11]   = 0x00
    # 4 calibration-ish bytes at write[12-15] correspond to read[14-17] — preserve verbatim.
    payload[12:16] = a[14:18]
    payload[16:18] = b"\x00\x00"
    payload[18:20] = a[20:22]            # humid_high_alarm — preserve
    payload[20]   = unit if unit is not None else a[22]
    payload[21:28] = encode_time(set_time if set_time is not None else datetime.datetime.now())
    payload[28]   = 0x00

    # Vendor software pattern: handshake op=0x05, then op=0x03
    request(h, bytes([0x05]), idle_ms=200, max_ms=1000)
    rs = request(h, bytes(payload), idle_ms=300, max_ms=2000)
    for r in rs:
        if r[0] == RID and r[1] == 0x03 and r[2:5] == b"\x00\x00\x03":
            return True
    raise RuntimeError(f"settings-write ack not seen: {[x.hex(' ') for x in rs]}")

def write_name(h, new_name: str):
    if len(new_name.encode("ascii")) > 16:
        raise ValueError("name must be <= 16 ASCII chars")
    payload = bytearray(53)
    payload[0]  = 0x12
    name_bytes = new_name.encode("ascii")
    payload[11:11 + len(name_bytes)] = name_bytes
    rs = request(h, bytes(payload), idle_ms=300, max_ms=2000)
    for r in rs:
        app = r[2:2 + r[1]]
        if r[0] == RID and len(app) >= 3 and app[:3] == b"\x00\x00\x12":
            return True
    raise RuntimeError(f"name-write ack not seen: {[x.hex(' ') for x in rs]}")

# ---------------------------------------------------------------- subcommands

def cmd_info(args):
    h = open_dev()
    try:
        m  = parse_meta(get_meta(h))
        nm = get_name(h)
        md = get_model(h)
        t, rh = get_live(h)
    finally:
        h.close()
    print(f"Model:           {md}")
    print(f"Logger name:     {nm}")
    print(f"Max points:      {m['max_points']}")
    print(f"Record count:    {m['record_count']}")
    print(f"Sample rate:     {fmt_duration(m['sample_sec'])}")
    print(f"Capacity:        {fmt_long_duration(m['max_points'] * m['sample_sec'])} "
          f"at this sample rate")
    print(f"LED flash cycle: {m['led_cycle_sec']} s")
    print(f"Start mode:      {'Manual' if m['start_mode'] else 'Instant'}")
    print(f"Temp unit:       {'F' if m['unit'] else 'C'}")
    print(f"Session start:   {m['start_time'].isoformat(sep=' ')}")
    if t is not None:
        print(f"Live reading:    {t:.2f} C / {rh:.2f} %RH")

def cmd_fetch(args):
    h = open_dev()
    try:
        m = parse_meta(get_meta(h))
        if m['record_count'] == 0:
            sys.exit("No records on device.")
        h.write(make_req(bytes([0x01])))
        reports = read_until_idle(h, idle_ms=400, max_ms=15000)
    finally:
        h.close()

    pages = {}
    for r in reports:
        if r[0] != RID: continue
        length = r[1]
        app = r[2:2 + length]
        if length == 3: continue   # trailer
        seq = (app[0] << 8) | app[1]
        pages[seq] = app[2:]

    raw = b"".join(d for _, d in sorted(pages.items()))
    if len(raw) % 4 or len(raw) // 4 != m['record_count']:
        print(f"WARNING: got {len(raw)} bytes ({len(raw)//4} records), "
              f"expected {m['record_count']}. Device may have run out of data, "
              f"or you've already fetched once on this power cycle (one-shot).",
              file=sys.stderr)

    interval = datetime.timedelta(seconds=m['sample_sec'])
    start    = m['start_time']
    n = len(raw) // 4

    rows = []
    for i in range(n):
        t_raw, rh_raw = struct.unpack_from("<hh", raw, i * 4)
        elapsed_sec = i * m['sample_sec']
        rows.append((start + i * interval,
                     elapsed_sec / 60.0,
                     elapsed_sec / 3600.0,
                     elapsed_sec / 86400.0,
                     t_raw / 100.0, rh_raw / 100.0))

    headers = ["timestamp", "elapsed_minutes", "elapsed_hours", "elapsed_days",
               "temperature_c", "humidity_rh"]

    # Use integer formatting for an elapsed unit only when every sample lands on
    # a whole boundary; otherwise show enough precision to distinguish samples.
    min_fmt  = "0"      if m['sample_sec'] %    60 == 0 else "0.00"
    hour_fmt = "0"      if m['sample_sec'] %  3600 == 0 else "0.0000"
    day_fmt  = "0"      if m['sample_sec'] % 86400 == 0 else "0.000000"

    if args.path.lower().endswith(".xlsx"):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "records"
        ws.append(headers)
        for r in rows:
            ws.append(list(r))
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 14
        for cell in ws["A"][1:]:
            cell.number_format = "yyyy-mm-dd hh:mm:ss"
        for cell in ws["B"][1:]:
            cell.number_format = min_fmt
        for cell in ws["C"][1:]:
            cell.number_format = hour_fmt
        for cell in ws["D"][1:]:
            cell.number_format = day_fmt
        for cell in ws["E"][1:] + ws["F"][1:]:
            cell.number_format = "0.00"
        wb.save(args.path)
    else:
        # Translate openpyxl formats to Python format specs.
        py_min  = ".0f" if min_fmt  == "0" else ".2f"
        py_hour = ".0f" if hour_fmt == "0" else ".4f"
        py_day  = ".0f" if day_fmt  == "0" else ".6f"
        with open(args.path, "w", newline="") as f:
            w = csv.writer(f)
            w.writerow(headers)
            for ts, em, eh, ed, t, rh in rows:
                w.writerow([ts.isoformat(sep=" "),
                            format(em, py_min),
                            format(eh, py_hour),
                            format(ed, py_day),
                            f"{t:.2f}", f"{rh:.2f}"])

    print(f"Wrote {n} records to {args.path}")
    print(f"  span: {start} ... {start + (n-1)*interval}  (every {m['sample_sec']}s)")

def cmd_set(args):
    if not any([args.led is not None, args.sample is not None,
                args.start, args.unit, args.name is not None, args.time]):
        sys.exit("nothing to set; pass --led/--sample/--start/--unit/--name/--time")

    h = open_dev()
    try:
        meta_app = get_meta(h)
        old_name = get_name(h)

        kwargs = {}
        if args.led    is not None: kwargs["led"]    = args.led
        if args.sample is not None: kwargs["sample"] = args.sample
        if args.start:              kwargs["start_mode"] = (1 if args.start == "manual" else 0)
        if args.unit:               kwargs["unit"]  = (1 if args.unit  == "f" else 0)
        if args.time:               kwargs["set_time"] = datetime.datetime.now()

        new_name = args.name if args.name is not None else old_name
        settings_changing = bool(kwargs)

        if settings_changing:
            print("WARNING: settings-write resets the recording session and erases "
                  "records. Run 'fetch' first if you need them.", file=sys.stderr)
            write_settings(h, meta_app, **kwargs)
            # op=0x03 wipes the logger name as a side effect — re-push it.
            write_name(h, new_name)
            print("Settings written.")
        elif args.name is not None:
            write_name(h, new_name)
            print(f"Name set to '{new_name}'.")

        # Read back and show
        new_meta = parse_meta(get_meta(h))
        nm = get_name(h)
        print(f"\nNow on device:")
        capacity = new_meta['max_points'] * new_meta['sample_sec']
        print(f"  name={nm!r}  sample={new_meta['sample_sec']}s  "
              f"led={new_meta['led_cycle_sec']}s  "
              f"start={'Manual' if new_meta['start_mode'] else 'Instant'}  "
              f"unit={'F' if new_meta['unit'] else 'C'}  "
              f"clock={new_meta['start_time']}")
        print(f"  capacity: {fmt_long_duration(capacity)} "
              f"({new_meta['max_points']} samples × {fmt_duration(new_meta['sample_sec'])})")
    finally:
        h.close()

# ---------------------------------------------------------------- entry

def main():
    p = argparse.ArgumentParser(prog="trh", description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    sp = p.add_subparsers(dest="cmd", required=True)

    sp.add_parser("info", help="show device info & live reading")

    pf = sp.add_parser("fetch", help="dump records to CSV")
    pf.add_argument("path")

    ps = sp.add_parser("set", help="change settings")
    ps.add_argument("--led",    type=parse_duration, metavar="N",
                    help="LED flash cycle, e.g. '5', '5s', '30s'")
    ps.add_argument("--sample", type=parse_duration, metavar="N",
                    help="sample rate, e.g. '4s', '60', '5m', '1h'")
    ps.add_argument("--start",  choices=["manual", "instant"])
    ps.add_argument("--unit",   choices=["c", "f"])
    ps.add_argument("--name",   metavar="STR", help="logger name (<=16 ASCII chars)")
    ps.add_argument("--time",   action="store_true", help="sync device clock to host")

    args = p.parse_args()
    {"info": cmd_info, "fetch": cmd_fetch, "set": cmd_set}[args.cmd](args)

if __name__ == "__main__":
    sys.stdout.reconfigure(line_buffering=True)
    main()
